from flask import Flask,request, jsonify , send_from_directory ,send_file
import os,sys,json
from flask.globals import session
import mysql.connector
from datetime import datetime
from openpyxl import  load_workbook
import openpyxl as xl

import requests
from requests.api import delete
from mockAPI import mocked_json, mocked_sessions

app = Flask(__name__)
upload_folder = 'in/'
xlfile='rates.xlsx'

# db contenction
def init_db():
    return mysql.connector.connect(
        host='db',
        database='db',
        user='root',
        password='root'
    )



@app.route('/',methods = ['GET'])
def index():
    # connect = init_db()  
    # cur = connect.cursor()  
    # cur.execute("SHOW TABLES;")
    # rows = cur.fetchall()
    # resp = jsonify(rows)
    # resp.status_code = 200
    # return resp 
    return "ok" 

#11111111111111111111111111111111111111
@app.route('/health',methods = ['GET'])
def health():
    try:
        connect = init_db()  
        # mycursor = connect.cursor()  
        # mycursor.execute("show tables")
        # res = str(mycursor.fetchall())
    except:
        return "failed connecting to the database", 500
    else:
        return "WELCOME DATA CONNECTION WORKS" ,200

#2222222222222222222222222222222222222
@app.route('/provider',methods = ['POST'])
def creat_provider():
    try:
        pro_name=request.args.get('name')
        conn = init_db()
        mycursor = conn.cursor()
        query = (f"INSERT INTO providers (providername) VALUES ('{pro_name}')")
        mycursor.execute(query)
        conn.commit()
    except:  
        return "Name already exists"
    else:
        mycursor.execute(f"SELECT * FROM providers WHERE providername = '{pro_name}';")
        rows = mycursor.fetchmany(size=1)
        resp = jsonify(rows)
        resp.status_code = 200
        return resp
#333333333333333333333333333333333333
@app.route("/rates", methods=['POST'])
def upload_xl_data():
    filename=upload_folder+request.args.get('filename')
    global xlfile
    xlfile = request.args.get('filename')
    connect = init_db()  
    cur = connect.cursor()  
    book = load_workbook(filename)
    sheet = book.active

    query = """REPLACE INTO products (product_name, rate, scope) VALUES (%s , %s, %s)"""

    for r in sheet.iter_rows(2, sheet.max_row):
        product_id = r[0].value
        rate = r[1].value
        scope = r[2].value
        values=(product_id,rate,scope)
        cur.execute(query,values)

    connect.commit()
    return "a" ,200
#444444444444444444444444444444444444444
@app.route("/rates",methods=['GET'])
def download_file():
    try:
        return send_from_directory(upload_folder, xlfile, as_attachment=True)
    except:
        return
#5555555555555555555555555555555555555555555
@app.route('/truck',methods = ['POST'])
def creat_truck():
    try:
        pro_id=request.args.get('providerid')
        pro_lic=request.args.get('truckid')
        conn = init_db()
        mycursor = conn.cursor()
        query = (f"INSERT INTO trucks (truckid,providerid) VALUES ('{pro_lic}','{pro_id}')")
        mycursor.execute(query)
        conn.commit()
        return 'ok'
    except:
        return "ProviderID not Found"
#6666666666666666666666666666666666666666666
@app.route('/truck/<id>',methods = ['PUT'])
def update_truckprovider(id):
    try:
        new_id = request.args.get('providerid')
        conn = init_db()
        mycursor = conn.cursor()
        query = (f"UPDATE trucks SET providerid = '{new_id}' WHERE truckid = '{id}'")
        mycursor.execute(query)
        conn.commit()
        return "OK"
    except:
        return "Invalid input"



@app.route('/provider/<id>',methods = ['PUT'])
def update_name(id):
    try:
        new_name=request.args.get('name')
        conn = init_db()
        mycursor = conn.cursor()
        query = (f"UPDATE providers SET providername = '{new_name}' WHERE id = '{id}'")
        mycursor.execute(query)
        conn.commit()
    except:
        return "Name already exists"
    else:   
        return "OK"
#7777777777777777777777777777777777777777

@app.route('/truck/<id>', methods=['GET'])
def itemId(id):
    now = datetime.now()
    time = now.strftime("%Y%m")
    test_id = id
    _from = request.args.get('from')
    _to = request.args.get('to')
    if not _to:
        _to = now.strftime("%Y%m%d%H%M%S")
        # _to=11111111111111
    if not _from:
        _from = time + '01000000'
        # _from=88888888888888
    conn = init_db()
    cursor = conn.cursor()
    # ty:
    int_id=int(id)
    query=f"SELECT DISTINCT neto FROM sessions WHERE date=(SELECT MAX(date) FROM sessions WHERE trucks_id={int_id});"
    cursor.execute(query) 
    netoCursor = cursor.fetchall()
    query_sessions=f"SELECT id FROM sessions WHERE trucks_id={int_id} AND date BETWEEN {_from} AND {_to};"
    cursor.execute(query_sessions) 
    rows=[] 
    rows = cursor.fetchall()
    session={
    "id":0,
    "tara":0,
    }
    if not rows:
        session["id"] = 404,
        session["tara"] = 'N/A' 
    else:
        session = { 
        "id":int(test_id),
        "tara":netoCursor[0],
        "sessions":[]
        }  
        for i in range(0, len(rows)):
            session["sessions"].append(rows[i])
    resp = jsonify(session)
    resp.status_code = 200
    return resp




        
@app.route("/weight" , methods=['GET'])
def get_weight():
    connect = init_db()
    cur = connect.cursor(dictionary=True, buffered=True)
    fromTime = request.args.get('from') if request.args.get('from') else datetime.now().strftime("%Y%m%d000000")
    toTime = request.args.get('to') if request.args.get('to') else datetime.now().strftime("%Y%m%d%H%M%S")
    filter = f"('{request.args.get('filter')}')" if request.args.get('filter') else "('in' , 'out' , 'none')" 
     
    query="""SELECT t1.id, direction, bruto, neto, product_name, GROUP_CONCAT(t3.containers_id) as containers 
    FROM sessions AS t1 JOIN products AS t2 ON t1.products_id = t2.id 
    JOIN containers_has_sessions as t3 ON t1.id = t3.sessions_id 
    WHERE t1.date BETWEEN '{0}' AND '{1}' AND direction IN {2} GROUP BY t3.sessions_id"""
    cur.execute(query.format(fromTime, toTime, filter))
    rows = cur.fetchall()
    return jsonify(rows)

#888888888888888888888888888888888888888888888

@app.route('/bill/<id>', methods=['GET'])
def billId(id):
    now = datetime.now()
    time = now.strftime("%Y%m")
    test_id = id
    _from = request.args.get('from')
    _to = request.args.get('to')
    if not _to:
        _to = now.strftime("%Y%m%d%H%M%S")
        # _to=11111111111111
    if not _from:
        _from = time + '01000000'
        # _from=88888888888888
    conn = init_db()
    cursor = conn.cursor() 
    cursor.execute(f'SELECT providername FROM providers WHERE id = "{id}"')
    n=cursor.fetchall()
    cursor.execute(f'SELECT id FROM trucks WHERE providerid = "{id}"')
    trucks=cursor.fetchall()


    sessionCount=0      
    getTrucks_list=[]
    session_list=[]
   
    for i in trucks:
        getTrucks_list.append(json.loads(mocked_json))
        
        # getTrucks_list.append(json.loads(requests.get('http://localhost:5000/truck/77777?from=11111111111111&to=88888888888888')))
    # return jsonify(getTrucks_list)                           
        
                                                       
    for dic in getTrucks_list:
        sessionCount += len(dic["sessions"])
        for i in dic["sessions"]:
            session_list.append(i)
        
    # return jsonify(session_list)

    session_response=[]
    neto_produce_list=[]
    for session in session_list:
        # session_response.append(json.loads(requests.get(url=('http://localhost:5000/weight?from=11111111111111&to=88888888888888&filter=out'))))
        session_response.append(json.loads(mocked_sessions))
        # return (jsonify(session_response))
    
        for data in session_response:
            neto_produce_list.append({"neto":data["neto"],"produce":data["produce"]})
    
    # return jsonify(neto_produce_list)

    cursor.execute('select id from products')
    product_name=cursor.fetchall()    
    products=[]
    # return jsonify(product_name)

    for name in product_name:
        count=0
        amount=0
        for d in neto_produce_list:
            if name[0] == d["produce"]:
                count +=1
                amount += d["neto"]
        products.append({"product":name[0],"count":count,"amount":amount,"pay":0,"rates":0})
        
    cursor.execute('select id,rate,scope from products')
    rates_list=cursor.fetchall()
    # return jsonify(rates_list)
    total=0


    for dict in products:
      for row in rates_list:
        # if row[0] == dict["product"] and row[2] == 'All':
            dict["rates"]=row[1]
            dict["pay"]=dict["rates"] * dict ["amount"]
            total += dict["pay"]
        # elif row[0] == dict["product"] and row[2] == n[0][0]:
        #     dict["rates"]=row[1]
        #     dict["pay"]=dict["rates"] * dict ["amount"]
        #     total += dict["pay"]
    for dict in products:
        if dict["amount"]==0:
            products.remove(dict)
    for dict in products:
        if dict["amount"]==0:
            products.remove(dict)
    provider = { 
        "id": int(id),
        "name": n[0][0],
        "from": _from,
        "to": _to,
        "truckCount":len(trucks),
        "sessionCount":sessionCount,
        "products": products,
        "Total": total,
        }  





    conn.commit() 
    # response = requests.get(f"http://3.70.209.94:8083/weight?from=11111111111111&to=88888888888888&filter=out")
    resp = jsonify(provider)
    resp.status_code = 200
    return resp






if __name__ == '__main__':
    app.run(host="0.0.0.0",debug = False)





    




